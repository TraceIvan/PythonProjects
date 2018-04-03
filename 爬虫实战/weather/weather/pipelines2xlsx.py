# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html
import time
import os.path
import openpyxl

class WeatherPipeline(object):
    def process_item(self, item, spider):
        today=time.strftime('%Y_%m_%d',time.localtime())
        filename=today+'.xlsx'
        sheet_title = "徐州" + today + "天气"
        if not os.path.isfile(filename):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_title
            ws.append(list(dict(item).keys()))
            ws.append(list(dict(item).values()))
            wb.save(filename=filename)
        else:
            wb = openpyxl.load_workbook(filename=filename)
            if wb.get_sheet_by_name(sheet_title):
                ws=wb.get_sheet_by_name(sheet_title)
                ws.append(list(dict(item).values()))
                wb.save(filename=filename)
            else:
                wb.create_sheet(title=sheet_title)
                ws = wb.get_sheet_by_name(sheet_title)
                ws.append(list(dict(item).values()))
                wb.save(filename=filename)
        return item
